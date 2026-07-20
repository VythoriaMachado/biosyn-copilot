import os
import json
import subprocess
from datetime import date, datetime, timedelta
from flask import Flask, render_template, jsonify, request, send_file
from config import DOWNLOADS_PATH
# Nuvem (Render) → SQLite | Local → Excel
if os.environ.get("RENDER"):
    from db_handler import (
        get_today_activities, get_all_data, save_checklist,
        get_weekly_data, get_managerial_data, propose_next_day_schedule,
        _get_weekday,
    )
else:
    from excel_handler import (
        get_today_activities, get_all_data, save_checklist,
        get_weekly_data, get_managerial_data, propose_next_day_schedule,
        _get_weekday,
    )
from outlook_handler import get_today_outlook_events, get_next_day_events, is_authenticated, _fetch_events as _fetch_full_events
from insights import generate_insights
from dashboard_generator import generate_weekly_html

app = Flask(__name__)


@app.route("/")
def index():
    return render_template("index.html")


# ── TODAY DASHBOARD ─────────────────────────────────────────────────────────

@app.route("/api/today")
def api_today():
    # Checklist usa sempre o dia anterior; dashboard usa hoje
    is_checklist = request.args.get("checklist") == "1"
    ics_url = request.args.get("ics_url") or None
    usuario = request.args.get("usuario") or "Usuário"

    force_date = request.args.get("force_date") or None
    if force_date:
        try:
            target = datetime.strptime(force_date, "%d/%m/%Y").date()
        except Exception:
            target = date.today() - timedelta(days=1)
    elif is_checklist:
        target = date.today() - timedelta(days=1)
        if target.weekday() >= 5:
            target = target - timedelta(days=target.weekday() - 4)
    else:
        target = date.today()

    all_activities = get_next_day_events(target_date=target, ics_url=ics_url) if is_checklist else get_today_outlook_events(ics_url=ics_url)

    if is_checklist:
        # get_next_day_events retorna formato resumido; expandir para formato completo
        all_activities = _fetch_full_events(target, ics_url)

    all_activities.sort(key=lambda x: x.get("horario_inicio", "00:00"))
    for a in all_activities:
        a["responsavel"] = usuario

    horas_planejadas = sum(a.get("tempo_previsto", 60) for a in all_activities) / 60
    reunioes = [a for a in all_activities if a.get("tipo") == "reuniao" or
                any(w in a.get("titulo", "").lower() for w in ["reunião", "meeting", "call", "sync", "alinhamento"])]

    # Buscar registros já salvos (usa o handler correto: SQLite na nuvem, Excel local)
    target_str = target.strftime("%d/%m/%Y")
    all_records = get_all_data()
    usuario_lower = usuario.strip().lower() if usuario else ""
    excel_dia = [r for r in all_records if r.get("data", "") == target_str
                 and (not usuario_lower or r.get("responsavel", "").strip().lower() == usuario_lower)]

    # Para o dashboard: pendências são do dia de hoje (filtradas por usuário)
    excel_hoje = [r for r in all_records if r.get("data", "") == date.today().strftime("%d/%m/%Y")
                  and (not usuario_lower or r.get("responsavel", "").strip().lower() == usuario_lower)]
    pendencias = [a for a in excel_hoje if a.get("status") in ["Parcial", "Não realizado"]]

    # Montar mapa de respostas já salvas (por título)
    respostas_salvas = {}
    for r in excel_dia:
        titulo = r.get("titulo", "").strip().lower()
        if titulo:
            respostas_salvas[titulo] = {
                "status":               r.get("status", ""),
                "houve_atraso":         r.get("houve_atraso", ""),
                "motivo_atraso":        r.get("motivo_atraso", ""),
                "tempo_executado":      r.get("tempo_executado", ""),
                "tempo_excedente":      r.get("tempo_excedente", ""),
                "reagendado":           r.get("reagendado", ""),
                "prioridade":           r.get("prioridade", ""),
                "atividade_extra":      r.get("atividade_extra", ""),
                "categoria_extra":      r.get("categoria_extra", ""),
                "nome_atividade_extra": r.get("nome_atividade_extra", ""),
                "tempo_extra_label":    r.get("tempo_extra_label", ""),
                "tempo_extra":          r.get("tempo_extra", 0),
                "solicitante_extra":    r.get("solicitante_extra", ""),
                "observacoes":          r.get("observacoes", ""),
                "pauta_reuniao":        r.get("pauta_reuniao", ""),
            }

    # Atividades do Excel do dia que não estão no Outlook
    titulos_outlook = {a.get("titulo","").strip().lower() for a in all_activities}
    extras_excel = [
        {**r, "origem": "Excel"} for r in excel_dia
        if r.get("titulo","").strip().lower() not in titulos_outlook
           and r.get("titulo","").strip()
    ]
    all_activities = all_activities + extras_excel

    dias_pt = ["Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira", "Sexta-feira", "Sábado", "Domingo"]

    return jsonify({
        "data":        target.strftime("%d/%m/%Y"),
        "dia_semana":  dias_pt[target.weekday()],
        "atividades":  all_activities,
        "respostas_salvas": respostas_salvas,
        "ja_preenchido": len(respostas_salvas) > 0,
        "stats": {
            "total_atividades": len(all_activities),
            "total_reunioes":   len(reunioes),
            "horas_planejadas": round(horas_planejadas, 1),
            "horas_livres":     round(max(0, 8 - horas_planejadas), 1),
            "pendencias":       len(pendencias),
        },
    })


# ── CHECKLIST ───────────────────────────────────────────────────────────────

@app.route("/api/checklist/save", methods=["POST"])
def api_save_checklist():
    data = request.get_json()
    entries = data.get("entries", [])
    if not entries:
        return jsonify({"error": "Nenhuma entrada recebida"}), 400
    try:
        save_checklist(entries)
        return jsonify({"success": True, "saved": len(entries)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── AUTH STATUS ─────────────────────────────────────────────────────────────

@app.route("/api/auth/status")
def api_auth_status():
    return jsonify({"authenticated": is_authenticated()})

@app.route("/api/auth/connect", methods=["POST"])
def api_auth_connect():
    try:
        from outlook_handler import _get_token
        token = _get_token()
        return jsonify({"success": bool(token)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ── WEEKLY DASHBOARD ────────────────────────────────────────────────────────

@app.route("/api/dashboard/weekly")
def api_weekly():
    usuario = request.args.get("usuario", "")
    stats = get_weekly_data(usuario=usuario)
    insights = generate_insights(days=7, usuario=usuario)
    return jsonify({**stats, "insights": insights})


@app.route("/api/dashboard/weekly/export")
def api_weekly_export():
    usuario = request.args.get("usuario", "")
    stats = get_weekly_data(usuario=usuario)
    insights = generate_insights(days=7, usuario=usuario)
    html_path = generate_weekly_html(stats, insights)
    if html_path and os.path.exists(html_path):
        subprocess.Popen(["start", "", html_path], shell=True)
        return jsonify({"success": True, "path": html_path})
    return jsonify({"error": "Falha ao gerar dashboard"}), 500


# ── MANAGERIAL PANEL ────────────────────────────────────────────────────────

@app.route("/api/managerial")
def api_managerial():
    period = request.args.get("period", "month")
    usuario = request.args.get("usuario", "")
    data = get_managerial_data(period=period, usuario=usuario)
    insights = generate_insights(days={"month": 30, "quarter": 90, "year": 365}.get(period, 30), usuario=usuario)
    return jsonify({**data, "insights": insights})


# ── PLANNING ────────────────────────────────────────────────────────────────

@app.route("/api/planning/next-day")
def api_next_day():
    ics_url = request.args.get("ics_url") or None
    next_events = get_next_day_events(ics_url=ics_url)
    proposal = propose_next_day_schedule(outlook_events=next_events)
    return jsonify(proposal)


@app.route("/api/planning/add-pending", methods=["POST"])
def api_add_pending():
    data = request.get_json()
    today = date.today()
    next_day = today + timedelta(days=1)
    if next_day.weekday() >= 5:
        next_day = today + timedelta(days=7 - today.weekday())
    dias_pt = ["Segunda-feira","Terça-feira","Quarta-feira","Quinta-feira","Sexta-feira","Sábado","Domingo"]
    try:
        save_checklist([{
            "data":           next_day.strftime("%d/%m/%Y"),
            "dia_semana":     dias_pt[next_day.weekday()],
            "titulo":         data.get("titulo", "Demanda pendente"),
            "horario_inicio": "",
            "horario_fim":    "",
            "tempo_previsto": 0,
            "descricao":      data.get("descricao", ""),
            "responsavel":    data.get("responsavel", ""),
            "origem":         "Reagendado",
            "status":         "Pendente",
            "prioridade":     data.get("prioridade", "Média"),
            "solicitante_extra": data.get("solicitante", ""),
            "categoria_extra":   data.get("departamento", ""),
        }])
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/planning/apply", methods=["POST"])
def api_apply_planning():
    return jsonify({"success": True})


# ── INSIGHTS ────────────────────────────────────────────────────────────────

@app.route("/api/insights")
def api_insights():
    days = int(request.args.get("days", 30))
    usuario = request.args.get("usuario", "")
    return jsonify(generate_insights(days=days, usuario=usuario))


# ── HISTORY ─────────────────────────────────────────────────────────────────

@app.route("/api/history")
def api_history():
    data = get_all_data()
    return jsonify({"records": data, "total": len(data)})


@app.route("/api/history/date")
def api_history_date():
    data_str = request.args.get("data", "")
    usuario  = request.args.get("usuario", "")
    all_records = get_all_data()
    records = [r for r in all_records if r.get("data", "") == data_str]
    if usuario:
        records = [r for r in records if r.get("responsavel", "").strip().lower() == usuario.strip().lower()]
    return jsonify({"records": records, "total": len(records)})


@app.route("/api/history/delete-date", methods=["POST"])
def api_history_delete_date():
    data = request.get_json()
    date_str = data.get("data", "")
    usuario  = data.get("usuario", "")
    if not date_str:
        return jsonify({"error": "Data não informada"}), 400
    try:
        from db_handler import delete_by_date_user
        delete_by_date_user(date_str, usuario)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/history/update", methods=["POST"])
def api_history_update():
    data = request.get_json()
    record_id  = data.get("id")
    status     = data.get("status", "")
    houve_atraso = data.get("houve_atraso", "")
    observacoes  = data.get("observacoes", "")
    if not record_id:
        return jsonify({"error": "ID não informado"}), 400
    try:
        from db_handler import update_checklist_entry
        update_checklist_entry(record_id, status, houve_atraso, observacoes)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── EXPORTAÇÃO GESTOR ────────────────────────────────────────────────────────

@app.route("/api/gestor/checklist/exportar")
def api_gestor_exportar():
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data_inicio = request.args.get("data_inicio", "")  # YYYY-MM-DD (vem do input date)
    data_fim    = request.args.get("data_fim", "")
    usuario     = request.args.get("usuario", "").strip().lower()

    def to_iso(data_str):
        """Converte DD/MM/YYYY ou YYYY-MM-DD para YYYY-MM-DD para comparação."""
        s = str(data_str).strip()
        if len(s) == 10 and s[2] == '/' and s[5] == '/':
            # DD/MM/YYYY → YYYY-MM-DD
            return f"{s[6:]}-{s[3:5]}-{s[:2]}"
        return s[:10]  # já está em YYYY-MM-DD ou ISO timestamp

    all_records = get_all_data()

    filtered = []
    for r in all_records:
        d_iso = to_iso(r.get("data", ""))
        if data_inicio and d_iso < data_inicio:
            continue
        if data_fim and d_iso > data_fim:
            continue
        if usuario and r.get("responsavel", "").strip().lower() != usuario:
            continue
        filtered.append(r)

    filtered.sort(key=lambda r: (
        to_iso(r.get("data", "")),
        r.get("responsavel", ""),
        r.get("horario_inicio", "")
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    navy_fill   = PatternFill("solid", fgColor="002468")
    sky_fill    = PatternFill("solid", fgColor="EFF6FF")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    normal_font = Font(size=10)
    bold_font   = Font(bold=True, size=10)
    thin_side   = Side(style="thin", color="D0D7E0")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        "Data", "Dia da Semana", "Responsável", "Atividade",
        "Horário Início", "Horário Fim", "Tempo Previsto (min)",
        "Status", "Tempo Executado", "Houve Atraso?", "Motivo Atraso",
        "Reagendado?", "Prioridade", "Observações",
        "Atividade Extra?", "Solicitante Extra", "Origem", "Pauta da Reunião"
    ]
    col_widths = [13, 16, 20, 42, 15, 12, 22, 15, 20, 14, 28, 14, 12, 35, 15, 22, 14, 50]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = header_font
        cell.fill      = navy_fill
        cell.alignment = center
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30

    status_label = {
        "concluido": "Concluído", "Concluído": "Concluído",
        "parcial": "Parcial", "Parcial": "Parcial",
        "nao_realizado": "Não realizado", "Não realizado": "Não realizado",
        "": "—"
    }

    # cores por status
    fill_concluido    = PatternFill("solid", fgColor="D4F5E5")
    fill_parcial      = PatternFill("solid", fgColor="FEF3D6")
    fill_nao_realizado= PatternFill("solid", fgColor="FFE8E8")
    fill_alt          = sky_fill

    for row_idx, r in enumerate(filtered, 2):
        status_raw = r.get("status", "")
        status_txt = status_label.get(status_raw, status_raw or "—")
        atraso = r.get("houve_atraso", "")
        atraso_txt = "Sim" if str(atraso).lower() in ("sim", "true", "1") else "Não"

        vals = [
            r.get("data", ""),
            r.get("dia_semana", ""),
            r.get("responsavel", ""),
            r.get("titulo", r.get("atividade", "")),
            r.get("horario_inicio", ""),
            r.get("horario_fim", ""),
            r.get("tempo_previsto", ""),
            status_txt,
            r.get("tempo_executado", ""),
            atraso_txt,
            r.get("motivo_atraso", ""),
            r.get("reagendado", ""),
            r.get("prioridade", ""),
            r.get("observacoes", ""),
            r.get("atividade_extra", ""),
            r.get("solicitante_extra", ""),
            r.get("origem", ""),
            r.get("pauta_reuniao", ""),
        ]

        # cor da linha por status
        if status_raw in ("concluido", "Concluído"):
            row_fill = fill_concluido
        elif status_raw in ("parcial", "Parcial"):
            row_fill = fill_parcial
        elif status_raw in ("nao_realizado", "Não realizado"):
            row_fill = fill_nao_realizado
        elif row_idx % 2 == 0:
            row_fill = fill_alt
        else:
            row_fill = None

        long_cols = {4, 11, 14}  # Atividade, Motivo Atraso, Observações
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(v) if v is not None else "")
            cell.font   = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                horizontal="left" if col_idx in long_cols else "center",
                wrap_text=(col_idx in long_cols)
            )
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_idx].height = 18

    # linha de totais
    row_t = len(filtered) + 2
    ws.cell(row=row_t, column=1, value="TOTAL DE REGISTROS").font = bold_font
    ws.cell(row=row_t, column=2, value=len(filtered)).font = Font(bold=True, size=10, color="002468")
    concluidos    = sum(1 for r in filtered if r.get("status","") in ("concluido","Concluído"))
    parciais      = sum(1 for r in filtered if r.get("status","") in ("parcial","Parcial"))
    nao_realizados= sum(1 for r in filtered if r.get("status","") in ("nao_realizado","Não realizado"))
    ws.cell(row=row_t, column=3, value=f"Concluídos: {concluidos} | Parcial: {parciais} | Não realizados: {nao_realizados}").font = normal_font

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    periodo = f"{data_inicio}_a_{data_fim}" if data_inicio and data_fim else "completo"
    nome = f"checklist_gestor_{periodo}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── DOCUMENTOS ───────────────────────────────────────────────────────────────

from documentos_handler import (
    list_documentos, get_documento, create_documento,
    delete_documento, get_categorias_doc, upload_documento, _tipo_arquivo,
)

@app.route("/api/documentos")
def api_list_documentos():
    search    = request.args.get("search", "")
    categoria = request.args.get("categoria", "")
    tipo      = request.args.get("tipo", "")
    docs = list_documentos(search=search, categoria=categoria, tipo=tipo)
    return jsonify({"documentos": docs, "total": len(docs)})

@app.route("/api/documentos/categorias")
def api_doc_categorias():
    return jsonify({"categorias": get_categorias_doc()})

@app.route("/api/documentos/upload", methods=["POST"])
def api_doc_upload():
    file      = request.files.get("file")
    nome      = request.form.get("nome", "").strip()
    categoria = request.form.get("categoria", "Geral").strip()
    descricao = request.form.get("descricao", "").strip()
    criado_por = request.form.get("criado_por", "").strip()

    if not file or not file.filename:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    nome_arquivo = nome or file.filename
    tipo = _tipo_arquivo(file.filename)

    # verifica extensão permitida
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    ALLOWED = {'pdf','png','jpg','jpeg','webp','gif','xls','xlsx','xlsm','csv','doc','docx'}
    if ext not in ALLOWED:
        return jsonify({"error": f"Extensão .{ext} não permitida"}), 400

    try:
        url = upload_documento(file)
        doc = create_documento({
            "nome":       nome_arquivo,
            "tipo":       tipo,
            "url":        url,
            "categoria":  categoria,
            "descricao":  descricao,
            "criado_por": criado_por,
            "extensao":   ext,
        })
        return jsonify({"success": True, "documento": doc})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/documentos/<int:doc_id>", methods=["DELETE"])
def api_delete_documento(doc_id):
    try:
        delete_documento(doc_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── GUIAS (COMO FAZER) ───────────────────────────────────────────────────────

from guias_handler import (
    list_guias, get_guia, get_guia_por_atividade, create_guia,
    update_guia, delete_guia, toggle_favorito, duplicar_guia,
    vincular_guia, get_versoes, get_categorias, upload_midia,
)

@app.route("/api/guias")
def api_list_guias():
    search       = request.args.get("search", "")
    categoria    = request.args.get("categoria", "")
    favoritos    = request.args.get("favoritos", "") == "1"
    try:
        guias = list_guias(search=search, categoria=categoria, favoritos_only=favoritos)
        return jsonify({"guias": guias, "total": len(guias)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/categorias")
def api_guias_categorias():
    try:
        return jsonify({"categorias": get_categorias()})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/por-atividade")
def api_guia_por_atividade():
    titulo = request.args.get("titulo", "")
    if not titulo:
        return jsonify({"guia": None})
    try:
        return jsonify({"guia": get_guia_por_atividade(titulo)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>")
def api_get_guia(guia_id):
    try:
        guia = get_guia(guia_id)
        if not guia:
            return jsonify({"error": "Não encontrado"}), 404
        return jsonify(guia)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias", methods=["POST"])
def api_create_guia():
    data = request.get_json()
    try:
        guia_id = create_guia(data, usuario=data.get("usuario", ""))
        return jsonify({"success": True, "id": guia_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>", methods=["PUT"])
def api_update_guia(guia_id):
    data = request.get_json()
    try:
        update_guia(guia_id, data, usuario=data.get("usuario", ""))
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>", methods=["DELETE"])
def api_delete_guia(guia_id):
    try:
        delete_guia(guia_id)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>/favorito", methods=["POST"])
def api_toggle_favorito(guia_id):
    try:
        novo = toggle_favorito(guia_id)
        return jsonify({"success": True, "favorito": novo})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>/duplicar", methods=["POST"])
def api_duplicar_guia(guia_id):
    data = request.get_json() or {}
    try:
        novo_id = duplicar_guia(guia_id, usuario=data.get("usuario", ""))
        return jsonify({"success": True, "id": novo_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>/vincular", methods=["POST"])
def api_vincular_guia(guia_id):
    data = request.get_json()
    titulo_atividade = data.get("titulo_atividade", "")
    if not titulo_atividade:
        return jsonify({"error": "Título não informado"}), 400
    try:
        vincular_guia(guia_id, titulo_atividade)
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/upload", methods=["POST"])
def api_guia_upload():
    if "file" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400
    file = request.files["file"]
    if not file or not file.filename:
        return jsonify({"error": "Arquivo inválido"}), 400
    try:
        url = upload_midia(file)
        return jsonify({"success": True, "url": url, "nome": file.filename})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/guias/<int:guia_id>/versoes")
def api_versoes_guia(guia_id):
    try:
        return jsonify({"versoes": get_versoes(guia_id)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/status")
def api_status():
    """Diagnóstico: testa conexão e retorna contagem de registros."""
    try:
        records = get_all_data()
        usuarios = list({r.get("responsavel","") for r in records if r.get("responsavel")})
        datas = sorted({r.get("data","") for r in records if r.get("data")}, reverse=True)[:5]
        return jsonify({
            "ok": True,
            "total_registros": len(records),
            "usuarios": usuarios,
            "ultimas_datas": datas,
            "banco": "Supabase" if os.environ.get("SUPABASE_URL") else "SQLite local",
        })
    except Exception as e:
        return jsonify({"ok": False, "erro": str(e)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(debug=False, host="0.0.0.0", port=port)
