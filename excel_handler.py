import os
import shutil
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from config import EXCEL_PATH, SHEET_CHECKLIST

# ── Estrutura real do Excel ────────────────────────────────────────────────
# Col 0: Data           | Col 1: Dia          | Col 2: Responsável
# Col 3: Ordem          | Col 4: Horário Plan. | Col 5: Atividade
# Col 6: Duração (min)  | Col 7: Status        | Col 8: ◀ CAPA (ignorar)
# Col 9: Tempo Restante | Col 10: Ativ. Alt.   | Col 11: Solicitante/Setor
# Col 12: Motivo        | Col 13: Prioridade   | Col 14: Reuniões
# Col 15: Ev. Extra.    | Col 16: Observações  | Col 17: Preenchido Atraso
# Col 18: Registrado em


def _normalize_date(val):
    """Converte qualquer formato de data para DD/MM/YYYY."""
    if val is None:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    # ISO: 2026-04-27
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%d/%m/%Y")
        except Exception:
            continue
    return s


def _parse_horario(h):
    """Extrai inicio e fim de '08:00-10:00'."""
    if not h:
        return "", ""
    h = str(h).strip()
    if "-" in h:
        parts = h.split("-")
        return parts[0].strip(), parts[1].strip()
    return h, ""


def _row_to_record(row):
    """Converte tupla do Excel para dicionário normalizado."""
    inicio, fim = _parse_horario(row[4] if len(row) > 4 else "")
    return {
        "data":               _normalize_date(row[0] if len(row) > 0 else None),
        "dia_semana":         str(row[1]) if len(row) > 1 and row[1] else "",
        "responsavel":        str(row[2]) if len(row) > 2 and row[2] else "",
        "horario_inicio":     inicio,
        "horario_fim":        fim,
        "titulo":             str(row[5]) if len(row) > 5 and row[5] else "",
        "tempo_previsto":     int(row[6]) if len(row) > 6 and row[6] and str(row[6]).isdigit() else 60,
        "status":             str(row[7]) if len(row) > 7 and row[7] else "",
        "tempo_executado":    str(row[9])  if len(row) > 9  and row[9]  else "",
        "nome_atividade_extra": str(row[10]) if len(row) > 10 and row[10] else "",
        "solicitante_extra":  str(row[11]) if len(row) > 11 and row[11] else "",
        "motivo_atraso":      str(row[12]) if len(row) > 12 and row[12] else "",
        "prioridade":         str(row[13]) if len(row) > 13 and row[13] else "Média",
        "observacoes":        str(row[16]) if len(row) > 16 and row[16] else "",
        "houve_atraso":       str(row[17]) if len(row) > 17 and row[17] else "Não",
        "timestamp_registro": str(row[18]) if len(row) > 18 and row[18] else "",
        "origem":             "Excel",
        "descricao":          "",
    }


def _load_workbook():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Planilha não encontrada: {EXCEL_PATH}")
    return openpyxl.load_workbook(EXCEL_PATH, keep_vba=True)


def _ensure_sheet(wb):
    return wb[SHEET_CHECKLIST]


def get_today_activities(usuario=None):
    try:
        wb = _load_workbook()
        ws = _ensure_sheet(wb)
        today_str = date.today().strftime("%d/%m/%Y")
        activities = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            if _normalize_date(row[0]) != today_str:
                continue
            rec = _row_to_record(row)
            if usuario and rec["responsavel"].strip().lower() != usuario.strip().lower():
                continue
            activities.append(rec)
        wb.close()
        return activities
    except Exception:
        return []


def get_all_data():
    try:
        wb = _load_workbook()
        ws = _ensure_sheet(wb)
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[5]:  # precisa ter data e atividade
                data.append(_row_to_record(row))
        wb.close()
        return data
    except Exception:
        return []


def save_checklist(entries):
    try:
        wb = _load_workbook()
        ws = _ensure_sheet(wb)
        timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        # Descobrir próxima ordem do dia
        today_str = date.today().strftime("%d/%m/%Y")
        ordem = 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and _normalize_date(row[0]) == today_str:
                ordem += 1

        for entry in entries:
            inicio = entry.get("horario_inicio", "")
            fim    = entry.get("horario_fim", "")
            horario = f"{inicio}-{fim}" if inicio and fim else inicio

            solicitante = entry.get("solicitante_extra", "")
            categoria   = entry.get("categoria_extra", "")
            sol_setor   = f"{solicitante} / {categoria}" if solicitante and categoria else (solicitante or categoria)

            obs_parts = []
            if entry.get("houve_atraso") == "Sim":
                obs_parts.append(f"Atraso: {entry.get('motivo_atraso','')}")
            if entry.get("reagendado") == "Sim":
                obs_parts.append("Reagendado")
            if entry.get("atividade_extra") == "Sim":
                obs_parts.append(f"Extra: {entry.get('nome_atividade_extra','')}")
            if entry.get("tempo_executado"):
                obs_parts.append(f"Tempo: {entry.get('tempo_executado')}")

            next_row = ws.max_row + 1
            # Escrever na mesma estrutura do Excel original (19 colunas)
            values = [
                today_str,                                # Col 0: Data
                entry.get("dia_semana", _get_weekday()),  # Col 1: Dia
                entry.get("responsavel", ""),             # Col 2: Responsável
                ordem,                                    # Col 3: Ordem
                horario,                                  # Col 4: Horário Planejado
                entry.get("titulo", ""),                  # Col 5: Atividade
                entry.get("tempo_previsto", 60),          # Col 6: Duração (min)
                entry.get("status", ""),                  # Col 7: Status
                None,                                     # Col 8: ◀ CAPA (vazio)
                entry.get("tempo_extra", ""),             # Col 9: Tempo Restante
                entry.get("nome_atividade_extra", ""),    # Col 10: Ativ. Alternativa
                sol_setor,                                # Col 11: Solicitante/Setor
                entry.get("motivo_atraso", ""),           # Col 12: Motivo
                entry.get("prioridade", "Média"),         # Col 13: Prioridade
                "",                                       # Col 14: Reuniões
                entry.get("origem", ""),                  # Col 15: Evento Extra
                " | ".join(obs_parts),                    # Col 16: Observações
                entry.get("houve_atraso", "Não"),         # Col 17: Preenchido Atraso
                timestamp,                                # Col 18: Registrado em
            ]

            for col, val in enumerate(values, 1):
                cell = ws.cell(row=next_row, column=col, value=val)
                cell.font = Font(name="Calibri", size=10)
                if next_row % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="EBF5FB")
                cell.border = Border(bottom=Side(style="thin", color="DDDDDD"))
                cell.alignment = Alignment(wrap_text=False)

            ordem += 1

        wb.save(EXCEL_PATH)
        wb.close()
        return True
    except Exception as e:
        raise RuntimeError(f"Erro ao salvar: {e}")


def get_weekly_data(reference_date=None):
    if reference_date is None:
        reference_date = date.today()
    start = reference_date - timedelta(days=reference_date.weekday())
    end   = start + timedelta(days=4)

    try:
        records = get_all_data()
        df = pd.DataFrame(records)
        if df.empty:
            return _empty_weekly()
        df["Data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
        week_df = df[(df["Data_dt"].dt.date >= start) & (df["Data_dt"].dt.date <= end)].copy()
        if week_df.empty:
            return _empty_weekly()
        return _compute_weekly_stats(week_df)
    except Exception:
        return _empty_weekly()


def get_historical_data(days=90):
    try:
        records = get_all_data()
        df = pd.DataFrame(records)
        if df.empty:
            return pd.DataFrame()
        df["Data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
        cutoff = date.today() - timedelta(days=days)
        return df[df["Data_dt"].dt.date >= cutoff]
    except Exception:
        return pd.DataFrame()


def _compute_weekly_stats(df):
    total        = len(df)
    concluidas   = len(df[df["status"] == "Concluído"])
    parciais     = len(df[df["status"] == "Parcial"])
    nao_real     = len(df[df["status"] == "Não realizado"])
    reunioes     = len(df[df["titulo"].str.contains("reunião|meeting|call", case=False, na=False)])
    horas_prev   = df["tempo_previsto"].apply(lambda x: int(x) if str(x).isdigit() else 60).sum() / 60
    taxa_concl   = round((concluidas / total) * 100, 1) if total > 0 else 0

    by_day = {}
    for _, row in df.iterrows():
        day = str(row.get("dia_semana", ""))
        if day not in by_day:
            by_day[day] = {"concluidas": 0, "total": 0}
        by_day[day]["total"] += 1
        if row.get("status") == "Concluído":
            by_day[day]["concluidas"] += 1

    motivos = df[df["motivo_atraso"] != ""]["motivo_atraso"].value_counts().to_dict() if "motivo_atraso" in df else {}

    return {
        "total": total, "concluidas": concluidas, "parciais": parciais,
        "nao_realizadas": nao_real, "reunioes": reunioes,
        "horas_previstas": round(horas_prev, 1), "horas_executadas": 0,
        "horas_extras": 0, "horas_pendentes": 0,
        "taxa_conclusao": taxa_concl, "taxa_produtividade": 0,
        "by_day": by_day, "motivos_atraso": motivos, "categorias_extras": {},
        "atividades_detail": df[["titulo","status","tempo_previsto","dia_semana"]].to_dict("records"),
    }


def _empty_weekly():
    return {
        "total": 0, "concluidas": 0, "parciais": 0, "nao_realizadas": 0,
        "reunioes": 0, "horas_previstas": 0, "horas_executadas": 0,
        "horas_extras": 0, "horas_pendentes": 0, "taxa_conclusao": 0,
        "taxa_produtividade": 0, "by_day": {}, "motivos_atraso": {},
        "categorias_extras": {}, "atividades_detail": [],
    }


def get_managerial_data(period="month"):
    days_map = {"month": 30, "quarter": 90, "year": 365}
    df = get_historical_data(days=days_map.get(period, 30))
    if df.empty:
        return {}

    total      = len(df)
    concluidas = len(df[df["status"] == "Concluído"])
    taxa       = round((concluidas / total) * 100, 1) if total > 0 else 0
    reunioes   = len(df[df["titulo"].str.contains("reunião|meeting|call", case=False, na=False)])

    heatmap = {d: 0 for d in ["Segunda-feira","Terça-feira","Quarta-feira","Quinta-feira","Sexta-feira"]}
    for _, row in df.iterrows():
        day = str(row.get("dia_semana", ""))
        if day in heatmap:
            heatmap[day] += 1

    top_activities = df.groupby("titulo").size().sort_values(ascending=False).head(10).to_dict()
    motivos = {}
    if "motivo_atraso" in df:
        motivos = df[df["motivo_atraso"].notna() & (df["motivo_atraso"] != "")]["motivo_atraso"].value_counts().head(10).to_dict()

    return {
        "period": period, "total_atividades": total,
        "concluidas": concluidas, "taxa_conclusao": taxa,
        "reunioes": reunioes, "extras_total": 0,
        "by_categoria": {}, "heatmap_days": heatmap,
        "top_activities": top_activities, "motivos_atraso": motivos,
    }


def propose_next_day_schedule(outlook_events=None):
    today    = date.today()
    next_day = today + timedelta(days=1)
    if next_day.weekday() >= 5:
        next_day = today + timedelta(days=(7 - today.weekday()))

    schedule = []
    if outlook_events:
        for ev in outlook_events:
            schedule.append({
                "titulo":         ev.get("titulo", "Evento"),
                "horario_inicio": _mins_to_time(ev.get("inicio", 8*60)),
                "horario_fim":    _mins_to_time(ev.get("fim", 9*60)),
                "tempo_previsto": ev.get("duracao", 60),
                "origem":         "Outlook",
                "prioridade":     "Alta",
                "tipo":           "evento",
            })

    used = sum(s["tempo_previsto"] for s in schedule)
    available = 10 * 60  # 8h úteis em minutos

    return {
        "data":              next_day.strftime("%d/%m/%Y"),
        "dia_semana":        _get_weekday(next_day),
        "schedule":          schedule,
        "horas_disponiveis": round(available / 60, 1),
        "horas_planejadas":  round(used / 60, 1),
        "backlog_min":       max(0, used - available),
        "ocupacao_pct":      round((used / available) * 100, 1) if available > 0 else 0,
    }


def _mins_to_time(mins):
    return f"{mins//60:02d}:{mins%60:02d}"


def _get_weekday(d=None):
    if d is None:
        d = date.today()
    days = ["Segunda-feira","Terça-feira","Quarta-feira","Quinta-feira","Sexta-feira","Sábado","Domingo"]
    return days[d.weekday()]
